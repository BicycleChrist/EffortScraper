import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink
import re
from typing import Dict, List, Tuple, Optional
import logging
import tempfile
from PIL import Image
from io import BytesIO
import os
import base64
from openpyxl.drawing.image import Image as XLImage

class CFECPermitScraper:
    def __init__(self, output_file: str = 'cfec_permits.xlsx'):
        """Initialize the scraper with output file configuration."""
        self.output_file = output_file
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        self.session = requests.Session()
        self.temp_files = []  # Track temp files

        # Configure logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)

    def _create_navigation_sheet(self, fishery_type: str) -> None:
        """Create a navigation sheet for a fishery type."""
        if fishery_type not in self.workbook.sheetnames:
            sheet = self.workbook.create_sheet(fishery_type)
            sheet['A1'] = f'{fishery_type} Permits Navigation'
            sheet['A1'].font = Font(bold=True, size=14)
            sheet['A2'] = 'Permit Code'
            sheet['B2'] = 'Region'
            sheet['C2'] = 'Gear Type'
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 25
            sheet.column_dimensions['C'].width = 20

    def _add_permit_to_navigation(self,
                                fishery_type: str,
                                permit_code: str,
                                region: str,
                                gear_type: str) -> None:
        """Add a permit entry to the navigation sheet with hyperlink."""
        sheet = self.workbook[fishery_type]
        next_row = sheet.max_row + 1

        # Add permit info
        sheet[f'A{next_row}'] = permit_code
        sheet[f'B{next_row}'] = region
        sheet[f'C{next_row}'] = gear_type

        # Add hyperlink to permit's data sheet
        sheet[f'A{next_row}'].hyperlink = f'#{permit_code}!A1'
        sheet[f'A{next_row}'].font = Font(color='0000FF', underline='single')

    def _clean_cell_data(self, cell_data: str) -> str:
        """Clean up cell data and standardize empty values."""
        # Trim whitespace
        cell_data = cell_data.strip()

        # List of values to consider as empty/null
        null_values = ['.', '-', '&nbsp;', 'N/A', 'NA', ' ']

        # Check if the cell contains only placeholder characters
        if cell_data in null_values or not cell_data:
            return ''

        # Handle currency values
        if cell_data.startswith('$'):
            try:
                # Remove $ and commas, convert to float
                value = float(cell_data.replace('$', '').replace(',', ''))
                return f"${value:,.2f}"
            except:
                pass

        return cell_data

    def _save_image_to_sheet(self, sheet, image_data: str, cell_location: str = 'A1') -> None:
        """Save a base64 encoded image to the Excel sheet."""
        try:
            self.logger.info("Processing base64 image data")

            if image_data.startswith('data:image'):
                base64_data = image_data.split(',')[1]
                image_binary = base64.b64decode(base64_data)

                # Create temp file
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                tmp.write(image_binary)
                tmp.flush()
                tmp_path = tmp.name

                # Track the temp file
                self.temp_files.append(tmp_path)

                self.logger.info(f"Image saved to temp file: {tmp_path}")

                # Add to sheet
                img_file = XLImage(tmp_path)
                img_file.width = 600
                img_file.height = 300

                self.logger.info(f"Adding image to sheet at position {cell_location}")
                sheet.add_image(img_file, cell_location)

        except Exception as e:
            self.logger.error(f"Error saving image: {str(e)}")
            self.logger.error("Full error details: ", exc_info=True)

    def _parse_permit_data(self, permit_code: str, html_content: str) -> Tuple[pd.DataFrame, List[str]]:
        """Parse the underlying data and extract images for a specific permit."""
        soup = BeautifulSoup(html_content, 'html.parser')

        # Extract images first
        image_data = []
        for div in soup.find_all('div', class_='section level3'):
            img_tag = div.find('img')
            if img_tag and 'src' in img_tag.attrs:
                image_data.append(img_tag['src'])

        # Parse table data
        headers = []
        for th in soup.find_all('th'):
            header_text = ' '.join(th.stripped_strings)
            headers.append(header_text)

        rows = []
        for tr in soup.find_all('tr'):
            row_data = []
            for td in tr.find_all('td'):
                cell_data = ' '.join(td.stripped_strings)
                cell_data = self._clean_cell_data(cell_data)
                row_data.append(cell_data)
            if row_data:
                rows.append(row_data)

        # Create DataFrame
        df = pd.DataFrame(rows)
        if len(df.columns) == len(headers):
            df.columns = headers

        # Clean DataFrame
        df = df.replace({
            '': pd.NA,
            '.': pd.NA,
            '-': pd.NA,
            ' ': pd.NA,
            '&nbsp;': pd.NA,
            'N/A': pd.NA,
            'NA': pd.NA,
        })

        # Convert numeric columns where appropriate
        for col in df.columns:
            if df[col].dtype == object:
                try:
                    numeric_series = pd.to_numeric(
                        df[col].str.replace('$', '').str.replace(',', ''),
                        errors='coerce'
                    )
                    if not numeric_series.isna().all():
                        df[col] = numeric_series
                except:
                    pass

        return df, image_data

    def _create_permit_sheet(self, permit_code: str, data: pd.DataFrame, image_data: List[str] = None) -> None:
        """Create a new sheet for permit data with images."""
        sheet_name = permit_code
        if sheet_name not in self.workbook.sheetnames:
            sheet = self.workbook.create_sheet(sheet_name)

            # Write DataFrame first
            data = data.replace({pd.NA: None})
            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    if pd.isna(value):
                        value = None
                    cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='CCCCCC',
                                              end_color='CCCCCC',
                                              fill_type='solid')

            # Add images to the right of the data if they exist
            if image_data:
                # Get the last column number from the DataFrame
                last_col = len(data.columns)
                # Start two columns over from the data
                image_col = last_col + 2

                for idx, img_data in enumerate(image_data):
                    # Convert column number to Excel column letter
                    col_letter = get_column_letter(image_col)
                    # Position each image with some horizontal spacing
                    cell_loc = f'{col_letter}1'  # Start at row 1
                    self._save_image_to_sheet(sheet, img_data, cell_loc)
                    # Move 10 columns right for next image if there is one
                    image_col += 10

    def _extract_region(self, link) -> str:
        """Extract region from table structure."""
        try:
            # Get the parent td cell
            cell = link.find_parent('td')
            # Get all cells in the header row that span this column
            header_row = cell.find_parent('table').find('tr')

            # Calculate which region this cell belongs to based on colspan
            column_index = 0
            current_cell = cell
            while current_cell.find_previous_sibling('td'):
                current_cell = current_cell.find_previous_sibling('td')
                column_index += 1

            # Find corresponding header by counting colspans
            current_span = 0
            for header in header_row.find_all('td', {'align': 'center', 'valign': 'bottom'}):
                colspan = int(header.get('colspan', 1))
                current_span += colspan
                if current_span > column_index:
                    # Extract region name from header
                    region_text = ' '.join(header.stripped_strings)
                    # Remove the leading code (e.g., "(A)") and clean up
                    region = re.sub(r'\([A-Z]+\)\s*', '', region_text).strip()
                    return region

            return "Unknown Region"
        except Exception as e:
            self.logger.error(f"Error extracting region: {str(e)}")
            return "Unknown Region"

    def _extract_gear_type(self, link) -> str:
        """Extract gear type from table structure."""
        try:
            # Find the row containing this permit
            row = link.find_parent('tr')
            # Get the gear type cell (second cell, after the empty spacer)
            gear_cell = row.find_all('td')[1]
            # Extract and clean the gear type text
            gear_type = gear_cell.get_text(strip=True)
            # Remove any leading/trailing whitespace and 'pre' tags if present
            gear_type = re.sub(r'</?pre>', '', gear_type).strip()
            return gear_type
        except Exception as e:
            self.logger.error(f"Error extracting gear type: {str(e)}")
            return "Unknown Gear Type"

    def _build_permit_url(self, base_url: str, href: str) -> str:
        """Build full URL for permit detail page."""
        # Get the base directory from the original URL
        base_dir = '/'.join(base_url.split('/')[:-1])
        # Join with the href
        return f"{base_dir}/{href}"

    def scrape_permit_page(self, url: str, fishery_type: str) -> None:
        """Scrape a main permit page and its underlying data."""
        try:
            response = self.session.get(url)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')

            # Create navigation sheet if it doesn't exist
            self._create_navigation_sheet(fishery_type)

            # Find and process all permit links
            for link in soup.find_all('a', href=re.compile(r'x_.*\.htm')):
                permit_code = link.get_text(strip=True)
                href = link['href']

                # Extract region and gear type from the table structure
                region = self._extract_region(link)
                gear_type = self._extract_gear_type(link)

                # Get permit details
                permit_url = self._build_permit_url(url, href)
                try:
                    detail_response = self.session.get(permit_url)
                    detail_response.raise_for_status()

                    # Parse permit data and create sheet
                    data_df, image_data = self._parse_permit_data(permit_code, detail_response.text)
                    self._create_permit_sheet(permit_code, data_df, image_data)

                    # Add to navigation
                    self._add_permit_to_navigation(fishery_type, permit_code, region, gear_type)

                    self.logger.info(f"Processed permit {permit_code} with {len(image_data)} images")

                except Exception as e:
                    self.logger.error(f"Error processing permit {permit_code}: {str(e)}")

        except Exception as e:
            self.logger.error(f"Error scraping {url}: {str(e)}")

    def save(self):
        """Save the workbook to file and clean up temp files."""
        try:
            self.workbook.save(self.output_file)
            self.logger.info(f"Saved data to {self.output_file}")
        finally:
            # Clean up temp files
            for tmp_file in self.temp_files:
                try:
                    os.unlink(tmp_file)
                except Exception as e:
                    self.logger.error(f"Error removing temp file {tmp_file}: {str(e)}")

    def scrape_all_fisheries(self):

        fishery_urls = {
            'Salmon': 'https://www.cfec.state.ak.us/pmtvalue/mnusalm.htm',
            'Herring': 'https://www.cfec.state.ak.us/bit/MNUHERR.htm',
            'Other Finfish': 'https://www.cfec.state.ak.us/bit/MNUOFIN.htm',
            'Crab': 'https://www.cfec.state.ak.us/bit/MNUCRAB.htm',
            'Other Species': 'https://www.cfec.state.ak.us/bit/MNUOTHR.htm'
        }

        for fishery_type, url in fishery_urls.items():
            self.logger.info(f"Starting scrape of {fishery_type}")
            self.scrape_permit_page(url, fishery_type)
            self.logger.info(f"Completed scrape of {fishery_type}")

        self.save()


if __name__ == "__main__":
    scraper = CFECPermitScraper('cfec_permits.xlsx')
    scraper.scrape_all_fisheries()
