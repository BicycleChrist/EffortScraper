import pandas as pd
import numpy as np
import xgboost as xgb
from math import radians, sin, cos, sqrt, atan2
from sklearn.preprocessing import LabelEncoder
import logging
from openpyxl.styles import PatternFill


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Port coordinates
# Reccent St Paul Island MDO prices can be obtained via the master rate schedule document
# Price of Marine Diesel = $6.18 per gallon plus taxes in July 2024; $7.19 in July 2022
# https://stpaulak.com/wp-content/uploads/2024/07CSP_2024MasterRateSchedule_UPDATED_22Jul24.pdf
# Due to the remoteness of both St paul and King cove, we can assume that prices in the less remote King cove will be at a slight discount to St paul
port_coordinates = {
    'Dillingham': (59.03865636911631, -158.4774533325954),
    'Yakutat': (59.56478746108611, -139.7412206054035),
    'Sand Point': (55.331788230225, -160.50359738863224),
    'Naknek': (58.72249702758417, -156.98980218678372),
    'Adak': (51.861794414113874, -176.6364490817307),
    'Kodiak': (57.78623279364595, -152.41299121692836),
    'Seward': (60.11629012805927, -149.43162457515695),
    'Homer': (59.60726724460732, -151.42504132671013),
    'Petersburg': (56.81167780114231, -132.96025492558314),
    'Wrangell': (56.46695186460168, -132.38364065828412),
    'Dutch Harbor': (53.89425864089317, -166.54323613689516),
    'Cordova': (60.547243970160615, -145.76788793049704),
    'Akutan': (54.12841559508686, -165.7778777590431),
    'Juneau': (58.38428553582909, -134.6458802501777),
    'Sitka': (57.05376183638209, -135.34949780572398),
    'Ketchikan': (55.343012147350635, -131.64854512244167),
    'St Paul Island': (57.78689854451251, -152.41106925000147),
    'King Cove': (55.05711552698481, -162.31944444367755)
}

# Column mapping
column_mapping = {
    'Dillingham_Avg': 'Dillingham',
    'Dillingham_Min': 'Dillingham',
    'Dillingham_Max': 'Dillingham',
    'Yakutat_Max': 'Yakutat',
    'Sand Point_Avg': 'Sand Point',
    'Sand Point_Max': 'Sand Point',
    'Naknek_Min': 'Naknek',
    'Adak_Avg': 'Adak',
    'Adak_Min': 'Adak',
    'Kodiak_Min': 'Kodiak',
    'Kodiak_Max': 'Kodiak',
    'Seward_Avg': 'Seward',
    'Homer_Avg': 'Homer',
    'Homer_Min': 'Homer',
    'Petersburg_Avg': 'Petersburg',
    'Petersburg_Max': 'Petersburg',
    'Wrangell_Min': 'Wrangell',
    'Wrangell_Max': 'Wrangell',
    'Dutch Harbor_Max': 'Dutch Harbor',
    'Cordova_Avg': 'Cordova',
    'Cordova_Min': 'Cordova',
    'Akutan_Min': 'Akutan',
    'Akutan_Max': 'Akutan',
    'Juneau_Min': 'Juneau',
    'Sitka_Avg': 'Sitka',
    'Sitka_Max': 'Sitka',
    'Ketchikan_Avg': 'Ketchikan',
    'St Paul Island_Avg': 'St Paul Island',
    'St Paul Island_Min': 'St Paul Island',
    'St Paul Island_Max': 'St Paul Island',
    'King Cove_Max': 'King Cove'
}

def haversine_distance(lat1, lon1, lat2, lon2):
    # Been told this needs to happen, not exacly sure if
    R = 6371
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    return R * c

def calculate_port_distance_matrix(port_coordinates):
    # distance matrix between ports
    ports = list(port_coordinates.keys())
    distances = pd.DataFrame(index=ports, columns=ports)

    for p1 in ports:
        for p2 in ports:
            lat1, lon1 = port_coordinates[p1]
            lat2, lon2 = port_coordinates[p2]
            distances.loc[p1, p2] = haversine_distance(lat1, lon1, lat2, lon2)

    return distances

def calculate_price_correlations(df, ports):
    # price correlations between all ports
    correlations = pd.DataFrame(index=ports, columns=ports)

    for p1 in ports:
        for p2 in ports:
            mask = df[p1].notna() & df[p2].notna()
            if mask.sum() >= 2:  # Need at least 2 points
                correlations.loc[p1, p2] = df[p1][mask].corr(df[p2][mask])
            else:
                correlations.loc[p1, p2] = 0

    return correlations

def get_nearby_prices(df, port, date):
    # +/- 3 month price reference
    nearby_prices = []
    for offset in range(-3, 4):  # -3 to +3 months
        try:
            offset_date = date + pd.DateOffset(months=offset)
            # Handle the case where we might get a Series instead of a single value
            price = df.loc[offset_date, port]
            if isinstance(price, pd.Series):
                price = price.iloc[0] if not price.empty else None

            if pd.notna(price):
                nearby_prices.append(price)
        except (KeyError, IndexError):
            continue
    return nearby_prices

def calculate_reasonable_bounds(prices, default_variation=0.3):
    # price variation threshholds
    if not prices:
        return None, None

    median = np.median(prices)
    std = np.std(prices) if len(prices) > 1 else median * default_variation

    # Allow larger variation for higher prices
    variation = max(default_variation, std / median) if median > 0 else default_variation

    return median * (1 - variation), median * (1 + variation)

def create_enhanced_features(df, target_port, port_distances, price_correlations):
    features = pd.DataFrame(index=df.index)

    # Base features
    features['Year'] = df.index.year
    features['Month'] = df.index.month
    features['Quarter'] = df.index.quarter
    features['IsSummer'] = (features['Month'] >= 6) & (features['Month'] <= 8)
    features['IsWinter'] = (features['Month'] <= 2) | (features['Month'] == 12)
    features['Month_Sin'] = np.sin(2 * np.pi * features['Month']/12)
    features['Month_Cos'] = np.cos(2 * np.pi * features['Month']/12)

    # Price history features
    features['Prev_Month_Price'] = df[target_port].shift(1)
    features['Prev_Quarter_Avg'] = df[target_port].rolling(window=3, min_periods=1).mean()

    # Special handling for remote ports
    if target_port in ['St Paul Island', 'King Cove']:
        if 'Dutch Harbor' in df.columns:
            features['Dutch_Harbor_Price'] = df['Dutch Harbor']
            features['Dutch_Harbor_Weight'] = 0.4

        if 'Sand Point' in df.columns:
            features['Sand_Point_Price'] = df['Sand Point']
            features['Sand_Point_Weight'] = 0.3

        if 'Dillingham' in df.columns:
            features['Dillingham_Price'] = df['Dillingham']
            features['Dillingham_Weight'] = 0.2
    else:
        # Regular port features
        closest_ports = port_distances[target_port].sort_values()
        closest_ports = closest_ports[closest_ports.index != target_port][:5]

        for i, (other_port, distance) in enumerate(closest_ports.items(), 1):
            if other_port in df.columns:
                corr = price_correlations.loc[target_port, other_port]
                weight = 1 / (distance + 1) * max(0.1, corr)
                features[f'Port_{i}_Price'] = df[other_port]
                features[f'Port_{i}_Weight'] = weight

    return features


def calculate_remote_port_premium(row, reference_port, min_premium, max_premium):
    """Calculate premium price based on reference port price"""
    if pd.isna(row[reference_port]):
        return None

    # Use random premium within range to add slight variation
    premium = np.random.uniform(min_premium, max_premium)
    return row[reference_port] * (1 + premium)

def interpolate_with_model(df, port, features, model_params=None):
    if model_params is None:
        model_params = {
            'n_estimators': 200,
            'learning_rate': 0.05,
            'max_depth': 8,
            'random_state': 42
        }

    train_mask = df[port].notna()
    pred_mask = df[port].isna()

    if not train_mask.any() or not pred_mask.any():
        return df[port].copy()

    result = df[port].copy()

    # Handle St Paul Island first
    if port == 'St Paul Island':
        temp_df = pd.DataFrame(index=df.index)

        if 'Dillingham' in df.columns:
            temp_df['dill_price'] = df.apply(
                lambda x: calculate_remote_port_premium(x, 'Dillingham', 0.01, 0.04),
                axis=1
            )

        if 'Dutch Harbor' in df.columns:
            temp_df['dutch_price'] = df.apply(
                lambda x: calculate_remote_port_premium(x, 'Dutch Harbor', 0.05, 0.08),
                axis=1
            )

        # Process each missing value
        for idx in df.index[pred_mask]:
            # Apply known fixed prices first
            known_prices = {
                pd.Timestamp('2024-07-01'): 6.18,
                pd.Timestamp('2022-07-01'): 7.19
            }

            if idx in known_prices:
                result.loc[idx] = known_prices[idx]
                continue

            premium_prices = temp_df.loc[idx].dropna()
            if not premium_prices.empty:
                result.loc[idx] = premium_prices.mean()
            else:
                X_pred = features.loc[[idx]].fillna(-999)
                if not X_pred.empty:
                    X_train = features[train_mask].fillna(-999)
                    y_train = df.loc[train_mask, port]
                    if not X_train.empty and not y_train.empty:
                        model = xgb.XGBRegressor(**model_params)
                        model.fit(X_train, y_train)
                        result.loc[idx] = model.predict(X_pred)[0]

    # Handle King Cove separately, after St Paul Island is complete
    elif port == 'King Cove':
        # Get St Paul Island prices - they should all be populated now
        st_paul_prices = df['St Paul Island']

        for idx in df.index[pred_mask]:
            st_paul_price = st_paul_prices.loc[idx]
            if pd.notna(st_paul_price):
                # Calculate King Cove as 2-5% discount from St Paul
                discount = np.random.uniform(0.02, 0.05)
                result.loc[idx] = st_paul_price * (1 - discount)
            else:
                # Fallback to Dillingham/Dutch Harbor reference if needed
                temp_prices = []

                if 'Dillingham' in df.columns and pd.notna(df.loc[idx, 'Dillingham']):
                    temp_prices.append(df.loc[idx, 'Dillingham'] * 1.20)  # 20% premium

                if 'Dutch Harbor' in df.columns and pd.notna(df.loc[idx, 'Dutch Harbor']):
                    temp_prices.append(df.loc[idx, 'Dutch Harbor'] * 1.15)  # 15% premium

                if temp_prices:
                    result.loc[idx] = np.mean(temp_prices)
                else:
                    # Model-based fallback as last resort
                    X_pred = features.loc[[idx]].fillna(-999)
                    if not X_pred.empty:
                        X_train = features[train_mask].fillna(-999)
                        y_train = df.loc[train_mask, port]
                        if not X_train.empty and not y_train.empty:
                            model = xgb.XGBRegressor(**model_params)
                            model.fit(X_train, y_train)
                            result.loc[idx] = model.predict(X_pred)[0]

    else:
        # Original model-based interpolation for non-remote ports remains unchanged
        X_train = features[train_mask].fillna(-999)
        y_train = df.loc[train_mask, port]

        model = xgb.XGBRegressor(**model_params)
        model.fit(X_train, y_train)

        for idx in df.index[pred_mask]:
            X_pred = features.loc[[idx]].fillna(-999)
            nearby_prices = get_nearby_prices(df, port, idx)
            min_bound, max_bound = calculate_reasonable_bounds(nearby_prices)

            pred = model.predict(X_pred)[0]

            if min_bound is not None and max_bound is not None:
                pred = np.clip(pred, min_bound, max_bound)

            result.loc[idx] = pred

    return result



def extend_predictions(df, num_months, port_distances, price_correlations):
    # try to make some big boy predictions
    # Create future dates
    last_date = df.index[-1]
    future_dates = pd.date_range(start=last_date + pd.DateOffset(months=1),
                               periods=num_months,
                               freq='MS')

    # extend DataFrame
    extended_df = df.copy()
    for date in future_dates:
        extended_df.loc[date] = np.nan

    # Make predictions for each port
    for port in df.columns:
        try:
            features = create_enhanced_features(extended_df, port, port_distances, price_correlations)
            extended_df[port] = interpolate_with_model(extended_df, port, features)
        except Exception as e:
            logging.error(f"Error predicting future values for {port}: {str(e)}")
            continue

    return extended_df




def preprocess_data(df, column_mapping):
    # very jenk should remove
    df_processed = pd.DataFrame(index=df.index)

    port_columns = {}
    for col, port in column_mapping.items():
        if port not in port_columns:
            port_columns[port] = []
        port_columns[port].append(col)

    for port, cols in port_columns.items():
        if len(cols) == 1:
            df_processed[port] = df[cols[0]]
        else:
            # Get the column with the most non-null values
            valid_counts = df[cols].notna().sum()
            best_col = valid_counts.idxmax()
            df_processed[port] = df[best_col]


            for col in cols:
                mask = df_processed[port].isna() & df[col].notna()
                df_processed.loc[mask, port] = df.loc[mask, col]

    return df_processed

def extend_predictions(df, num_months, port_distances, price_correlations):
    """Extend the dataset with predictions for future months"""
    # Create future dates
    last_date = df.index[-1]
    future_dates = pd.date_range(start=last_date + pd.DateOffset(months=1),
                               periods=num_months,
                               freq='MS')

    # Create extended DataFrame
    extended_df = df.copy()
    for date in future_dates:
        extended_df.loc[date] = np.nan

    # Make predictions for each port
    for port in df.columns:
        try:
            features = create_enhanced_features(extended_df, port, port_distances, price_correlations)
            extended_df[port] = interpolate_with_model(extended_df, port, features)
        except Exception as e:
            logging.error(f"Error predicting future values for {port}: {str(e)}")
            continue

    return extended_df


def adjust_premium_port_prices(df):
    """
    Adjusts prices for St Paul Island and King Cove based on remoteness premium
    """
    df_adjusted = df.copy()

    # Known recent prices for St Paul Island
    st_paul_reference = {
        pd.Timestamp('2024-07-01'): 6.18,  # From rate schedule
        pd.Timestamp('2022-07-01'): 7.19   # Historical reference
    }

    # Process each date
    for date in df_adjusted.index:
        # Find the highest non-premium port price for this date
        other_ports = df_adjusted.loc[date].drop(['St Paul Island', 'King Cove'])
        base_price = other_ports.max()

        if pd.notna(base_price):
            # St Paul Island adjustment
            if pd.isna(df_adjusted.loc[date, 'St Paul Island']):
                if date in st_paul_reference:
                    df_adjusted.loc[date, 'St Paul Island'] = st_paul_reference[date]
                else:
                    df_adjusted.loc[date, 'St Paul Island'] = base_price * 1.35

            # King Cove adjustment
            if pd.isna(df_adjusted.loc[date, 'King Cove']):
                df_adjusted.loc[date, 'King Cove'] = base_price * 1.25

    return df_adjusted





def main():
    try:
        logging.info("Loading and preprocessing data...")
        df = pd.read_csv('IITcsv.csv')

        if df.empty:
            raise ValueError("Input DataFrame is empty")

        df['Date'] = pd.to_datetime(df[['Year', 'Month']].assign(Day=1))
        df.set_index('Date', inplace=True)

        # Convert price columns to numeric
        for col in df.columns:
            if col in column_mapping:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace('$', '').str.replace(',', ''),
                                      errors='coerce')

        # Preprocess to get one column per port
        df_processed = preprocess_data(df, column_mapping)

        logging.info("Calculating port relationships...")
        port_distances = calculate_port_distance_matrix(port_coordinates)
        price_correlations = calculate_price_correlations(df_processed, port_coordinates.keys())

        logging.info("Interpolating missing values...")
        result_df = df_processed.copy()

        for port in port_coordinates.keys():
            if port not in result_df.columns:
                logging.warning(f"Skipping {port} - not found in data")
                continue

            try:
                logging.info(f"Processing {port}...")
                features = create_enhanced_features(result_df, port, port_distances, price_correlations)
                result_df[port] = interpolate_with_model(result_df, port, features)
            except Exception as e:
                logging.error(f"Error processing {port}: {str(e)}")
                continue

        # Save results with colored cells for interpolated values
        output_filename = 'interpolated_fuel_prices_final.xlsx'
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name='Prices')

            # Color interpolated cells
            worksheet = writer.sheets['Prices']
            interpolated_fill = PatternFill(start_color='ADD8E6',
                                          end_color='ADD8E6',
                                          fill_type='solid')

            # Color cells where values were interpolated (where original was NaN)
            for col_idx, port in enumerate(result_df.columns, start=2):
                for row_idx, (orig, interp) in enumerate(zip(df_processed[port], result_df[port]), start=2):
                    if pd.isna(orig) and not pd.isna(interp):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = interpolated_fill

        # Print validation statistics
        logging.info("\nValidation Statistics:")
        for port in port_coordinates.keys():
            if port in result_df.columns:
                original = df_processed[port].dropna()
                interpolated = result_df[port][df_processed[port].isna()]

                if not original.empty:
                    logging.info(f"\n{port}:")
                    logging.info(f"Original data range: ${original.min():.2f} - ${original.max():.2f}")
                    logging.info(f"Original mean: ${original.mean():.2f}")

                if not interpolated.empty:
                    logging.info(f"Interpolated data range: ${interpolated.min():.2f} - ${interpolated.max():.2f}")
                    logging.info(f"Interpolated mean: ${interpolated.mean():.2f}")
                    logging.info(f"Number of interpolated values: {len(interpolated)}")

        logging.info(f"\nResults saved to {output_filename}")

    except Exception as e:
        logging.error(f"An error occurred in main execution: {str(e)}")
        raise

if __name__ == "__main__":
    main()
